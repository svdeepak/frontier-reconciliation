"""XLSX exception report (M6).

Turns one case's agent output plus its source rows into the artefact an
operations analyst actually works from — three sheets:

  Summary     — run identity, row/match counts, break counts by type, and the
                aggregate-vs-row-level contrast that makes offsetting breaks
                visible even when totals tie out.
  Exceptions  — one row per break: type, row IDs, amounts, computed
                difference, confidence, suggested action.
  Evidence    — one row per cited source row, under its break, with the full
                source values behind the computed difference.

Deterministic: no LLM, no network. Amounts are written as numbers so the
sheet stays sortable and footable; identifiers stay text.

Usage:
  python -m recon.report --outputs outputs/solution --data data --out reports
"""

import argparse
import json
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from recon.agent import normalize

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13)
LABEL_FONT = Font(bold=True)
BREAK_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0.00"

SUMMARY_HEADERS = ["Metric", "Value"]
EXCEPTION_HEADERS = ["Break ID", "Type", "Component causes", "A row IDs", "B row IDs",
                     "Amount A", "Amount B", "Difference", "Confidence",
                     "Suggested action", "Evidence"]
EVIDENCE_HEADERS = ["Break ID", "Break type", "Source", "Row ID", "Date", "Reference",
                    "Currency", "FX rate", "Foreign ccy", "Foreign amount",
                    "Gross", "Fee", "Net"]


def _style_header(ws, headers: list[str], row: int = 1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, max_width: int = 60):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                longest = max(longest, min(len(str(v)), max_width))
        ws.column_dimensions[letter].width = max(10, min(longest + 2, max_width))


def _totals(rows) -> dict:
    return {
        "gross": sum((r.gross_amount for r in rows), Decimal("0")),
        "fee": sum((r.fee_amount for r in rows), Decimal("0")),
        "net": sum((r.net_amount for r in rows), Decimal("0")),
    }


def build_summary(ws, output: dict, a_rows, b_rows):
    ws.title = "Summary"
    ws["A1"] = f"Reconciliation exception report — {output['case_id']}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    breaks = output.get("breaks", [])
    by_type: dict[str, int] = {}
    for b in breaks:
        by_type[b["break_type"]] = by_type.get(b["break_type"], 0) + 1
    ta, tb = _totals(a_rows), _totals(b_rows)
    sysinfo = output.get("system", {})

    rows: list[tuple] = [
        ("Case ID", output["case_id"]),
        ("System", sysinfo.get("name", "")),
        ("Model", sysinfo.get("model", "")),
        ("Provider", sysinfo.get("provider", "")),
        ("Schema version", output.get("schema_version", "")),
        ("", ""),
        ("Source A rows", len(a_rows)),
        ("Source B rows", len(b_rows)),
        ("Clean matches", len(output.get("matches", []))),
        ("Exceptions (breaks)", len(breaks)),
        ("", ""),
        ("Source A gross total", float(ta["gross"])),
        ("Source B gross total", float(tb["gross"])),
        ("Gross difference (A-B)", float(ta["gross"] - tb["gross"])),
        ("Source A net total", float(ta["net"])),
        ("Source B net total", float(tb["net"])),
        ("Net difference (A-B)", float(ta["net"] - tb["net"])),
        ("Aggregate totals tie out", "YES" if ta["net"] == tb["net"] else "NO"),
        ("Row-level exceptions found", len(breaks)),
        ("Aggregate check misleading",
         "YES — totals tie out but row-level breaks exist"
         if ta["net"] == tb["net"] and breaks else "no"),
        ("", ""),
    ]
    for t in sorted(by_type):
        rows.append((f"Breaks — {t}", by_type[t]))

    _style_header(ws, SUMMARY_HEADERS, row=3)
    r = 4
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        if isinstance(value, float):
            c.number_format = MONEY
        if label.startswith("Aggregate check misleading") and str(value).startswith("YES"):
            c.font = Font(bold=True, color="9C0006")
        r += 1
    _autosize(ws)


def build_exceptions(ws, output: dict):
    _style_header(ws, EXCEPTION_HEADERS)
    for i, b in enumerate(output.get("breaks", []), start=2):
        vals = [
            b["break_id"], b["break_type"],
            ", ".join(b.get("component_causes", [])),
            ", ".join(b.get("a_row_ids", [])),
            ", ".join(b.get("b_row_ids", [])),
            b.get("amount_a"), b.get("amount_b"), b.get("difference"),
            b.get("confidence", ""), b.get("suggested_action", ""), b.get("evidence", ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=col >= 10)
            if col in (6, 7, 8) and isinstance(v, (int, float)):
                c.number_format = MONEY
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEPTION_HEADERS))}{max(1, ws.max_row)}"
    _autosize(ws, max_width=48)


def build_evidence(ws, output: dict, a_by_id: dict, b_by_id: dict):
    _style_header(ws, EVIDENCE_HEADERS)
    r = 2
    for b in output.get("breaks", []):
        for src, ids, table in (("A", b.get("a_row_ids", []), a_by_id),
                                ("B", b.get("b_row_ids", []), b_by_id)):
            for rid in ids:
                row = table.get(rid)
                vals = [b["break_id"], b["break_type"], src, rid]
                if row is None:
                    # Cannot happen for verified output; surfaced rather than hidden.
                    vals += ["ROW NOT FOUND IN SOURCE"] + [None] * 8
                else:
                    vals += [row.date, row.reference, row.currency,
                             float(row.fx_rate) if row.fx_rate is not None else None,
                             row.foreign_currency,
                             float(row.foreign_amount) if row.foreign_amount is not None else None,
                             float(row.gross_amount), float(row.fee_amount), float(row.net_amount)]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=r, column=col, value=v)
                    c.border = BORDER
                    if col == 1:
                        c.fill = BREAK_FILL
                    if col in (11, 12, 13) and isinstance(v, float):
                        c.number_format = MONEY
                    if col == 10 and isinstance(v, float):
                        c.number_format = MONEY
                r += 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EVIDENCE_HEADERS))}{max(1, ws.max_row)}"
    _autosize(ws, max_width=40)


def build_workbook(output: dict, a_rows, b_rows) -> Workbook:
    wb = Workbook()
    a_by_id = {r.row_id: r for r in a_rows}
    b_by_id = {r.row_id: r for r in b_rows}
    build_summary(wb.active, output, a_rows, b_rows)
    build_exceptions(wb.create_sheet("Exceptions"), output)
    build_evidence(wb.create_sheet("Evidence"), output, a_by_id, b_by_id)
    return wb


def write_case_report(output_json: Path, data_dir: Path, out_dir: Path) -> Path:
    output = json.loads(output_json.read_text())
    case_dir = data_dir / "cases" / output["case_id"]
    a_rows = normalize(case_dir / "source_a.csv")
    b_rows = normalize(case_dir / "source_b.csv")
    wb = build_workbook(output, a_rows, b_rows)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{output['case_id']}.xlsx"
    wb.save(path)
    return path


def run(outputs_dir: Path, data_dir: Path, out_dir: Path) -> list[Path]:
    manifest = json.loads((data_dir / "manifest.json").read_text())
    written = []
    for c in sorted(manifest["cases"], key=lambda x: x["case_id"]):
        p = outputs_dir / f"{c['case_id']}.json"
        if p.exists():
            written.append(write_case_report(p, data_dir, out_dir))
    return written


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--outputs", type=str, default="outputs/solution")
    p.add_argument("--data", type=str, default="data")
    p.add_argument("--out", type=str, default="reports")
    args = p.parse_args()
    paths = run(Path(args.outputs), Path(args.data), Path(args.out))
    print(f"wrote {len(paths)} XLSX report(s) -> {args.out}/")

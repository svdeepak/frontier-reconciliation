"""M6 XLSX report tests — deterministic, no credentials.

Verifies sheet structure and that the numbers written into the workbook are
the numbers from the verified output and the source rows.
"""

import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from recon import agent, report  # noqa: E402
from recon.providers import MockProvider  # noqa: E402

DATA = ROOT / "data"
SIG = "case_13_signature_adversarial"


@pytest.fixture()
def sig_book(tmp_path):
    out, _ = agent.solve_case(MockProvider(), DATA, SIG)
    (tmp_path / f"{SIG}.json").write_text(json.dumps(out))
    path = report.write_case_report(tmp_path / f"{SIG}.json", DATA, tmp_path / "reports")
    return load_workbook(path), out


def _summary(ws) -> dict:
    return {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)}


def test_three_named_sheets(sig_book):
    wb, _ = sig_book
    assert wb.sheetnames == ["Summary", "Exceptions", "Evidence"]


def test_summary_reports_row_and_break_counts(sig_book):
    wb, out = sig_book
    s = _summary(wb["Summary"])
    assert s["Case ID"] == SIG
    assert s["Source A rows"] == 13
    assert s["Source B rows"] == 13
    assert s["Exceptions (breaks)"] == len(out["breaks"]) == 3
    assert s["Clean matches"] == len(out["matches"])


def test_summary_flags_the_aggregate_trap(sig_book):
    """Totals tie out on this case; the summary must say so AND warn."""
    wb, _ = sig_book
    s = _summary(wb["Summary"])
    assert s["Net difference (A-B)"] == 0
    assert s["Aggregate totals tie out"] == "YES"
    assert s["Row-level exceptions found"] == 3
    assert str(s["Aggregate check misleading"]).startswith("YES")


def test_summary_counts_breaks_by_type(sig_book):
    wb, _ = sig_book
    s = _summary(wb["Summary"])
    assert s["Breaks — COMPOUND"] == 1
    assert s["Breaks — DUPLICATE"] == 1
    assert s["Breaks — MISSING_IN_B"] == 1


def test_exceptions_sheet_one_row_per_break_with_computed_difference(sig_book):
    wb, out = sig_book
    ws = wb["Exceptions"]
    assert [c.value for c in ws[1]] == report.EXCEPTION_HEADERS
    assert ws.max_row == 1 + len(out["breaks"])
    rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    for b in out["breaks"]:
        r = rows[b["break_id"]]
        assert ws.cell(row=r, column=2).value == b["break_type"]
        assert ws.cell(row=r, column=8).value == b["difference"]
        assert ws.cell(row=r, column=9).value == b["confidence"]
        assert ws.cell(row=r, column=10).value == b["suggested_action"]


def test_compound_components_are_listed(sig_book):
    wb, _ = sig_book
    ws = wb["Exceptions"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "COMPOUND":
            assert ws.cell(row=r, column=3).value == "FEE_MISMATCH, FX_DIFFERENCE"
            assert ws.cell(row=r, column=8).value == 0.61
            return
    pytest.fail("no COMPOUND row found")


def test_evidence_sheet_carries_source_values_for_every_cited_row(sig_book):
    wb, out = sig_book
    ws = wb["Evidence"]
    assert [c.value for c in ws[1]] == report.EVIDENCE_HEADERS
    expected = sum(len(b["a_row_ids"]) + len(b["b_row_ids"]) for b in out["breaks"])
    assert ws.max_row == 1 + expected

    a_by_id = {r.row_id: r for r in agent.normalize(DATA / "cases" / SIG / "source_a.csv")}
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value != "A":
            continue
        rid = ws.cell(row=r, column=4).value
        src = a_by_id[rid]
        assert ws.cell(row=r, column=5).value == src.date
        assert ws.cell(row=r, column=6).value == src.reference
        assert ws.cell(row=r, column=11).value == float(src.gross_amount)
        assert ws.cell(row=r, column=13).value == float(src.net_amount)


def test_fx_rows_carry_foreign_amount_and_rate(sig_book):
    """The FX component is only auditable if rate and foreign amount are shown."""
    wb, _ = sig_book
    ws = wb["Evidence"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=4).value == "A1011":
            assert ws.cell(row=r, column=8).value == 1.0857
            assert ws.cell(row=r, column=9).value == "EUR"
            assert ws.cell(row=r, column=10).value == 1234.5
            return
    pytest.fail("A1011 not present in Evidence sheet")


def test_amounts_written_as_numbers_not_text(sig_book):
    wb, _ = sig_book
    ws = wb["Exceptions"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=8).value
        assert isinstance(v, (int, float)), f"difference not numeric: {v!r}"


def test_clean_case_report_has_no_exception_rows(tmp_path):
    out, _ = agent.solve_case(MockProvider(), DATA, "case_01_clean_small")
    (tmp_path / "c.json").write_text(json.dumps(out))
    wb = load_workbook(report.write_case_report(tmp_path / "c.json", DATA, tmp_path / "r"))
    assert wb["Exceptions"].max_row == 1
    assert wb["Evidence"].max_row == 1
    s = _summary(wb["Summary"])
    assert s["Exceptions (breaks)"] == 0
    assert s["Aggregate check misleading"] == "no"


def test_run_writes_a_report_for_every_case(tmp_path):
    outputs = tmp_path / "solution"
    agent.run(DATA, outputs)
    paths = report.run(outputs, DATA, tmp_path / "reports")
    assert len(paths) == 14
    for p in paths:
        assert p.exists() and p.stat().st_size > 0
